# 报告生成模块

import html
from datetime import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ReportGenerator:
    """生成各种格式的评估报告"""
    
    @staticmethod
    def generate_html_report(filename, findings, stats, risk_score, risk_level):
        """生成 HTML 报告"""
        
        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>招标风险评估报告</title>
    <style>
        * {{ margin: 0; padding: 0; }}
        body {{ font-family: '微软雅黑', Arial; background: #f5f5f5; padding: 20px; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        h1 {{ color: #333; margin-bottom: 30px; text-align: center; font-size: 28px; }}
        .header-info {{ display: flex; justify-content: space-between; margin-bottom: 30px; font-size: 14px; color: #666; }}
        
        .score-section {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 8px; margin-bottom: 30px; text-align: center; }}
        .score-value {{ font-size: 60px; font-weight: bold; margin: 20px 0; }}
        .score-label {{ font-size: 20px; margin-bottom: 10px; }}
        .score-desc {{ font-size: 14px; opacity: 0.9; }}
        
        .stats-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr 1fr; gap: 15px; margin-bottom: 30px; }}
        .stat-card {{ background: #f9f9f9; padding: 20px; border-radius: 6px; border-left: 4px solid #667eea; }}
        .stat-label {{ font-size: 12px; color: #999; margin-bottom: 8px; }}
        .stat-value {{ font-size: 24px; font-weight: bold; color: #333; }}
        
        .findings-section {{ margin-bottom: 30px; }}
        .findings-section h2 {{ font-size: 18px; color: #333; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 2px solid #667eea; }}
        
        .risk-item {{ background: #f9f9f9; padding: 15px; margin-bottom: 15px; border-left: 4px solid #ccc; border-radius: 4px; }}
        .risk-item.level-a {{ border-left-color: #ff4444; }}
        .risk-item.level-b {{ border-left-color: #ff9800; }}
        .risk-item.level-c {{ border-left-color: #ffb74d; }}
        .risk-item.level-d {{ border-left-color: #4caf50; }}
        
        .risk-header {{ display: flex; justify-content: space-between; margin-bottom: 10px; }}
        .risk-id {{ font-weight: bold; color: #333; }}
        .risk-level {{ padding: 4px 12px; border-radius: 4px; font-size: 12px; font-weight: bold; color: white; }}
        .risk-level.a {{ background-color: #ff4444; }}
        .risk-level.b {{ background-color: #ff9800; }}
        .risk-level.c {{ background-color: #ffb74d; }}
        .risk-level.d {{ background-color: #4caf50; }}
        
        .risk-name {{ font-size: 14px; color: #333; margin-bottom: 8px; }}
        .risk-desc {{ font-size: 12px; color: #666; margin-bottom: 8px; }}
        .risk-keywords {{ font-size: 12px; color: #999; }}
        .risk-location-list {{ font-size: 12px; color: #666; margin: 4px 0 0 18px; }}
        .risk-recommendation {{ background: white; padding: 10px; margin-top: 10px; border-left: 3px solid #667eea; }}
        .rec-label {{ font-size: 12px; color: #667eea; font-weight: bold; }}
        .rec-text {{ font-size: 12px; color: #333; margin-top: 5px; }}
        
        .module-group {{ margin-bottom: 25px; }}
        .module-title {{ background: #f0f0f0; padding: 10px 15px; border-radius: 4px; font-weight: bold; color: #333; margin-bottom: 10px; }}
        
        .footer {{ text-align: center; color: #999; font-size: 12px; margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🎯 招标投标风险评估报告</h1>
        
        <div class="header-info">
            <div><strong>评估文件:</strong> {filename}</div>
            <div><strong>评估时间:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
            <div><strong>评估模型:</strong> v1.4（101 项规则）</div>
        </div>
        
        <div class="score-section">
            <div class="score-label">综合风险评分</div>
            <div class="score-value">{risk_score}/100</div>
            <div class="score-label">{risk_level}</div>
            <div class="score-desc">
                {ReportGenerator._get_score_advice(risk_score)}
            </div>
        </div>
        
        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-label">发现风险总数</div>
                <div class="stat-value">{stats['total']}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">A 级重大违规</div>
                <div class="stat-value" style="color: #ff4444;">{stats['by_level'].get('A', 0)}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">B 级严重违规</div>
                <div class="stat-value" style="color: #ff9800;">{stats['by_level'].get('B', 0)}</div>
            </div>
            <div class="stat-card">
                <div class="stat-label">高置信度发现</div>
                <div class="stat-value" style="color: #667eea;">{len(stats['high_confidence'])}</div>
            </div>
        </div>
        
        {ReportGenerator._generate_findings_html(findings, stats)}
        
        <div class="footer">
            <p>本报告由招标投标风险评估系统自动生成</p>
            <p>仅供参考，具体风险判断请参考完整规则库进行人工复核</p>
        </div>
    </div>
</body>
</html>
        """
        
        return html_content
    
    @staticmethod
    def generate_excel_report(filename, findings, stats, risk_score, risk_level):
        """生成 Excel 报告"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "评估结果"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 45
        ws.column_dimensions['J'].width = 45
        
        # 标题
        ws['A1'] = '招标投标风险评估报告'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # 基本信息
        ws['A3'] = '评估信息'
        ws['A3'].font = Font(bold=True)
        ws['A4'] = '文件名'
        ws['B4'] = filename
        ws['A5'] = '评估时间'
        ws['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A6'] = '综合评分'
        ws['B6'] = f"{risk_score}/100"
        ws['A7'] = '风险等级'
        ws['B7'] = risk_level
        
        # 统计信息
        ws['A9'] = '统计信息'
        ws['A9'].font = Font(bold=True)
        ws['A10'] = '发现总数'
        ws['B10'] = stats['total']
        ws['A11'] = 'A 级风险'
        ws['B11'] = stats['by_level'].get('A', 0)
        ws['A12'] = 'B 级风险'
        ws['B12'] = stats['by_level'].get('B', 0)
        ws['A13'] = '高置信度'
        ws['B13'] = len(stats['high_confidence'])
        
        # 详细发现
        ws['A15'] = '详细发现'
        ws['A15'].font = Font(bold=True, size=12)
        
        # 表头
        row = 17
        headers = ['规则ID', '模块', '风险名称', '风险等级', '置信度', '匹配关键词', '整改建议', '文档地址', '发现位置', '依据']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # 填充数据
        row = 18
        for finding in sorted(findings, key=lambda x: x['level']):
            ws.cell(row=row, column=1, value=finding['rule_id'])
            ws.cell(row=row, column=2, value=finding['module'])
            ws.cell(row=row, column=3, value=finding['name'])
            
            level_cell = ws.cell(row=row, column=4, value=finding['level'])
            if finding['level'] == 'A':
                level_cell.fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
                level_cell.font = Font(color="FFFFFF", bold=True)
            elif finding['level'] == 'B':
                level_cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
                level_cell.font = Font(color="FFFFFF", bold=True)
            
            ws.cell(row=row, column=5, value=f"{finding['confidence']:.0%}")
            ws.cell(row=row, column=6, value=', '.join(finding['matched_keywords'][:3]))
            ws.cell(row=row, column=7, value=finding['recommendation'])
            ws.cell(row=row, column=8, value=finding.get('source', ''))
            ws.cell(row=row, column=9, value=ReportGenerator._format_locations_text(finding.get('locations', [])))
            ws.cell(row=row, column=10, value=finding.get('legal_basis', ''))
            
            # 设置行高以适应长文本
            ws.row_dimensions[row].height = 30
            
            row += 1
        
        return wb
    
    @staticmethod
    def generate_text_report(filename, findings, stats, risk_score, risk_level):
        """生成文本报告"""
        
        report = f"""
╔═══════════════════════════════════════════════════════════════╗
║          招标投标风险评估报告                                    ║
╚═══════════════════════════════════════════════════════════════╝

📋 评估基本信息
{'─' * 60}
文件名称: {filename}
评估时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
评估模型: 招标风险评估系统 v1.4
统计规则: 101 项指标（7 个通用模块 + 北京市工程建设项目招标投标负面行为清单）

📊 综合风险评分
{'─' * 60}
综合评分: {risk_score}/100
风险等级: {risk_level}
{ReportGenerator._get_score_advice(risk_score)}

📈 风险统计
{'─' * 60}
发现总数:        {stats['total']} 项
├─ A 级 (重大违规):   {stats['by_level'].get('A', 0)} 项  ❌
├─ B 级 (严重违规):   {stats['by_level'].get('B', 0)} 项  ⚠️
├─ C 级 (中等风险):   {stats['by_level'].get('C', 0)} 项  ⚡
└─ D 级 (低风险):     {stats['by_level'].get('D', 0)} 项  ✓

置信度统计:
├─ 高置信度 (≥70%):   {len(stats['high_confidence'])} 项
└─ 低置信度 (<70%):   {len(stats['low_confidence'])} 项

{'─' * 60}

🔍 详细发现
{'─' * 60}
"""
        
        # 按模块分组
        by_module = {}
        for finding in findings:
            module = finding['module']
            if module not in by_module:
                by_module[module] = []
            by_module[module].append(finding)
        
        for module in sorted(by_module.keys()):
            findings_in_module = by_module[module]
            report += f"\n【{module}】 - 发现 {len(findings_in_module)} 项\n"
            
            for finding in sorted(findings_in_module, key=lambda x: x['level']):
                level_icon = '❌' if finding['level'] == 'A' else '⚠️' if finding['level'] == 'B' else '⚡' if finding['level'] == 'C' else '✓'
                locations_text = ReportGenerator._format_locations_text(finding.get('locations', []))
                report += f"""
  {level_icon} {finding['rule_id']} [{finding['level']}级] {finding['name']}
     描述: {finding['description']}
     文档: {finding.get('source', '')}
     匹配: {', '.join(finding['matched_keywords'][:3])}{'...' if len(finding['matched_keywords']) > 3 else ''}
     位置: {locations_text}
     置信: {finding['confidence']:.0%}
     {'依据: ' + finding['legal_basis'] if finding.get('legal_basis') else ''}
     💡  {finding['recommendation']}
"""
        
        report += f"""

📋 建议行动清单
{'─' * 60}

【立即处理（A 级风险）】 - {stats['by_level'].get('A', 0)} 项
"""
        
        a_level = [f for f in findings if f['level'] == 'A']
        for i, finding in enumerate(a_level, 1):
            report += f"\n{i}. {finding['rule_id']} - {finding['name']}\n"
            report += f"   整改: {finding['recommendation']}\n"
        
        if stats['by_level'].get('B', 0) > 0:
            report += f"""

【限期改正（B 级风险）】 - {stats['by_level'].get('B', 0)} 项 (建议 60 日内完成)
"""
            
            b_level = [f for f in findings if f['level'] == 'B']
            for i, finding in enumerate(b_level, 1):
                report += f"\n{i}. {finding['rule_id']} - {finding['name']}\n"
                report += f"   整改: {finding['recommendation']}\n"
        
        report += f"""

⚙️ 技术说明
{'─' * 60}
本报告由自动化评估系统生成，采用关键词匹配方式检测潜在风险。
为确保准确性，建议：
1. 人工复核高风险项（A 级和高置信度项）
2. 对低置信度项进行二次确认
3. 参考完整规则库进行详细分析
4. 咨询专业法律顾问进行最终判断

免责声明
{'─' * 60}
本评估仅供参考，不构成法律建议。
使用者应对自己的判断承担全部责任。

────────────────────────────────────────────────────────────────
生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
系统版本: v1.0
"""
        
        return report
    
    @staticmethod
    def _generate_findings_html(findings, stats):
        """生成发现部分的 HTML"""
        
        html_str = '<div class="findings-section"><h2>🔍 详细风险发现</h2>'

        # 按模块分组
        by_module = {}
        for finding in findings:
            module = finding['module']
            if module not in by_module:
                by_module[module] = []
            by_module[module].append(finding)

        for module in sorted(by_module.keys()):
            findings_in_module = sorted(by_module[module], key=lambda x: x['level'])
            html_str += f'<div class="module-group"><div class="module-title">{module} - 发现 {len(findings_in_module)} 项</div>'

            for finding in findings_in_module:
                level_lower = finding['level'].lower()
                html_str += f'''
<div class="risk-item level-{level_lower}">
    <div class="risk-header">
        <div class="risk-id">{finding['rule_id']}: {finding['name']}</div>
        <div class="risk-level {level_lower}">{finding['level']}级</div>
    </div>
    <div class="risk-desc">{finding['description']}</div>
    <div class="risk-keywords"><strong>匹配关键词:</strong> {', '.join(finding['matched_keywords'][:5])}{'...' if len(finding['matched_keywords']) > 5 else ''}</div>
    <div class="risk-keywords"><strong>文档地址:</strong> {html.escape(finding.get('source') or '')}</div>
    <div class="risk-locations"><strong>发现位置:</strong>{ReportGenerator._generate_locations_html(finding.get('locations', []))}</div>
    <div class="risk-keywords"><strong>置信度:</strong> {finding['confidence']:.0%}</div>
    {f'<div class="risk-keywords"><strong>依据:</strong> {html.escape(finding["legal_basis"])}</div>' if finding.get('legal_basis') else ''}
    <div class="risk-recommendation">
        <div class="rec-label">💡 整改建议</div>
        <div class="rec-text">{finding['recommendation']}</div>
    </div>
</div>
'''
            
            html_str += '</div>'

        html_str += '</div>'
        return html_str
    
    @staticmethod
    def _format_locations_text(locations):
        """生成单条风险发现的位置摘要文本（用于纯文本报告）"""

        if not locations:
            return '未定位到具体位置'

        parts = [
            f"{loc.get('location', '')}(……{loc.get('context', '')}……)"
            for loc in locations[:5]
        ]
        suffix = ' 等' if len(locations) > 5 else ''
        return '; '.join(parts) + suffix

    @staticmethod
    def _generate_locations_html(locations):
        """生成单条风险发现的位置列表 HTML（位置信息来自上传文档内容，必须转义防 XSS）"""

        if not locations:
            return '<ul class="risk-location-list"><li>未定位到具体位置</li></ul>'

        items = ''
        for loc in locations[:10]:
            location = html.escape(loc.get('location', ''))
            keyword = html.escape(loc.get('keyword', ''))
            context = html.escape(loc.get('context', ''))
            items += f'<li>{location}（关键词: {keyword}）：……{context}……</li>'

        return f'<ul class="risk-location-list">{items}</ul>'

    @staticmethod
    def _get_score_advice(score):
        """根据分数获取建议"""
        if score >= 80:
            return "✓ 风险较低，基本符合规范，继续监控"
        elif score >= 60:
            return "⚠️  中等风险，建议重点关注 B 级风险，逐步改进"
        elif score >= 40:
            return "❌ 高风险，必须立即整改 A 级风险，制定详细改正计划"
        else:
            return "🚨 严重风险，存在重大违规行为，需要紧急处理"
