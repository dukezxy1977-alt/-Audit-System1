import json
import os
import re
import sqlite3
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from flask import Flask, flash, redirect, render_template, request, send_from_directory, url_for
from werkzeug.utils import secure_filename

try:
    import docx
except Exception:
    docx = None

try:
    import pdfplumber
except Exception:
    pdfplumber = None


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.environ.get("DATA_DIR", BASE_DIR / "data"))
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", BASE_DIR / "uploads"))
DB_PATH = Path(os.environ.get("DATABASE_PATH", DATA_DIR / "audit_engine.db"))
DEFAULT_RULEBOOK = Path(os.environ.get("RULEBOOK_PATH", DATA_DIR / "工程项目招标投标审计规则引擎基础库_V10.xlsx"))

ALLOWED_RULE_EXT = {".xlsx"}
ALLOWED_DOC_EXT = {".txt", ".pdf", ".docx", ".doc"}
CORE_SHEETS = [
    "Rule_Master",
    "Decision_Master",
    "Evidence_Master",
    "Finding_Master",
    "Audit_Test_Master",
    "Audit_Logic_Master",
    "Audit_Question_Master",
    "Finding_Generator_Master",
    "Working_Paper_Generator",
    "Report_Generator",
    "Risk_Scoring_Master",
    "Knowledge_Graph_Master",
    "AI_Audit_Assistant_IO",
    "质量检查",
]

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-audit-rule-engine")
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("MAX_CONTENT_LENGTH", str(80 * 1024 * 1024)))
DATA_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH.parent.mkdir(parents=True, exist_ok=True)


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def init_db():
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS rulebooks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                filename TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS rule_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rulebook_id INTEGER NOT NULL,
                sheet_name TEXT NOT NULL,
                row_index INTEGER NOT NULL,
                rule_code TEXT,
                row_json TEXT NOT NULL,
                FOREIGN KEY(rulebook_id) REFERENCES rulebooks(id)
            );
            CREATE INDEX IF NOT EXISTS idx_rule_rows_sheet ON rule_rows(rulebook_id, sheet_name);
            CREATE INDEX IF NOT EXISTS idx_rule_rows_code ON rule_rows(rulebook_id, rule_code);

            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                description TEXT,
                created_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS uploaded_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                original_name TEXT NOT NULL,
                stored_name TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                text_content TEXT,
                uploaded_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id)
            );
            CREATE TABLE IF NOT EXISTS risk_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                rulebook_id INTEGER NOT NULL,
                rule_code TEXT NOT NULL,
                risk_level TEXT,
                score INTEGER,
                matched_terms TEXT,
                finding TEXT,
                legal_basis TEXT,
                working_paper_draft TEXT,
                report_paragraph TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id),
                FOREIGN KEY(rulebook_id) REFERENCES rulebooks(id)
            );
            """
        )


def active_rulebook():
    with db() as conn:
        return conn.execute("SELECT * FROM rulebooks WHERE active=1 ORDER BY id DESC LIMIT 1").fetchone()


def safe_cell(value):
    return "" if value is None else str(value).strip()


def infer_rule_code(sheet_name, record):
    if "二级规则编码" in record:
        return record.get("二级规则编码", "")
    if sheet_name == "Knowledge_Graph_Master":
        return record.get("二级规则编码", "")
    if sheet_name == "AI_Audit_Assistant_IO":
        return record.get("对应规则编码", "")
    return ""


def import_rulebook(xlsx_path, name=None):
    workbook_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    with db() as conn:
        conn.execute("UPDATE rulebooks SET active=0")
        cur = conn.execute(
            "INSERT INTO rulebooks(name, filename, imported_at, active) VALUES (?, ?, ?, 1)",
            (name or workbook_path.name, workbook_path.name, now()),
        )
        rulebook_id = cur.lastrowid

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [safe_cell(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
            if not any(headers):
                continue
            for row_idx in range(2, ws.max_row + 1):
                record = {}
                has_value = False
                for col_idx, header in enumerate(headers, start=1):
                    if not header:
                        continue
                    value = safe_cell(ws.cell(row_idx, col_idx).value)
                    record[header] = value
                    has_value = has_value or bool(value)
                if not has_value:
                    continue
                code = infer_rule_code(sheet_name, record)
                conn.execute(
                    "INSERT INTO rule_rows(rulebook_id, sheet_name, row_index, rule_code, row_json) VALUES (?, ?, ?, ?, ?)",
                    (rulebook_id, sheet_name, row_idx, code, json.dumps(record, ensure_ascii=False)),
                )
    return rulebook_id


def ensure_default_rulebook():
    init_db()
    if active_rulebook() is None and DEFAULT_RULEBOOK.exists():
        import_rulebook(DEFAULT_RULEBOOK, "工程项目招标投标审计规则引擎基础库_V10")


def sheet_rows(rulebook_id, sheet_name):
    with db() as conn:
        rows = conn.execute(
            "SELECT row_index, rule_code, row_json FROM rule_rows WHERE rulebook_id=? AND sheet_name=? ORDER BY row_index",
            (rulebook_id, sheet_name),
        ).fetchall()
    return [dict(row_index=r["row_index"], rule_code=r["rule_code"], data=json.loads(r["row_json"])) for r in rows]


def row_by_code(rulebook_id, sheet_name, code):
    with db() as conn:
        row = conn.execute(
            "SELECT row_json FROM rule_rows WHERE rulebook_id=? AND sheet_name=? AND rule_code=? ORDER BY row_index LIMIT 1",
            (rulebook_id, sheet_name, code),
        ).fetchone()
    return json.loads(row["row_json"]) if row else {}


def extract_text(path):
    ext = Path(path).suffix.lower()
    if ext == ".txt":
        return Path(path).read_text(encoding="utf-8", errors="ignore")
    if ext == ".docx" and docx:
        document = docx.Document(path)
        return "\n".join(p.text for p in document.paragraphs)
    if ext == ".pdf" and pdfplumber:
        chunks = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                chunks.append(page.extract_text() or "")
        return "\n".join(chunks)
    return f"文件名：{Path(path).name}"


def keywords_from_text(text):
    cleaned = re.sub(r"[，。；;、,.()\[\]（）《》“”\"'\s]+", " ", text)
    tokens = [w.strip() for w in re.split(r"\s+", cleaned) if len(w.strip()) >= 2]
    terms = []
    for token in tokens:
        terms.append(token)
        for part in re.findall(r"[\u4e00-\u9fff]{2,}", token):
            max_size = min(8, len(part))
            for size in range(max_size, 1, -1):
                for start in range(0, len(part) - size + 1):
                    terms.append(part[start:start + size])
    return list(dict.fromkeys(terms))


def build_match_terms(rule, logic, project_text):
    candidates = []
    for field in ["违规情景", "负面行为"]:
        candidates.extend(keywords_from_text(rule.get(field, "")))
    for index in range(1, 5):
        candidates.extend(keywords_from_text(logic.get(f"判断条件{index}", "")))

    stop_words = {
        "项目", "招标", "投标", "文件", "资料", "规定", "依法", "是否", "存在", "进行",
        "相关", "核实", "检查", "根据", "判断", "情形", "情况", "活动", "工程", "建设",
    }
    matched = []
    for term in candidates:
        if term in stop_words or len(term) < 3:
            continue
        if term in project_text and term not in matched:
            matched.append(term)
    return matched[:12]


def score_result(match_count, risk_level):
    base = {"高": 70, "中": 50, "较低": 25, "低风险": 25, "中风险": 50, "高风险": 70, "重大风险": 90}.get(risk_level, 40)
    return min(100, base + min(match_count * 3, 20))


def generate_risks(project_id):
    rb = active_rulebook()
    if not rb:
        raise RuntimeError("请先导入规则库 Excel")

    with db() as conn:
        files = conn.execute("SELECT * FROM uploaded_files WHERE project_id=?", (project_id,)).fetchall()
        conn.execute("DELETE FROM risk_results WHERE project_id=?", (project_id,))

    project_text = "\n".join(f["text_content"] or f["original_name"] for f in files)
    if not project_text.strip():
        project_text = "未上传可识别文本资料"

    created = 0
    rules = sheet_rows(rb["id"], "Rule_Master")
    with db() as conn:
        for item in rules:
            code = item["rule_code"]
            rule = item["data"]
            logic = row_by_code(rb["id"], "Audit_Logic_Master", code)
            finding = row_by_code(rb["id"], "Finding_Generator_Master", code)
            paper = row_by_code(rb["id"], "Working_Paper_Generator", code)
            report = row_by_code(rb["id"], "Report_Generator", code)
            terms = build_match_terms(rule, logic, project_text)
            if len(terms) < 2:
                continue

            risk_level = finding.get("风险等级") or logic.get("风险等级") or report.get("风险等级") or "中"
            score = score_result(len(terms), risk_level)
            finding_text = finding.get("审计发现模板") or report.get("审计发现模板") or f"发现疑似风险：{rule.get('违规情景', '')}"
            paper_text = "\n".join(
                [
                    f"审计事项：{paper.get('审计事项', '')}",
                    f"审计目标：{paper.get('审计目标', '')}",
                    f"审计程序：{paper.get('审计程序', '')}",
                    f"获取资料：{paper.get('获取资料', '')}",
                    f"审计发现：{paper.get('审计发现模板', finding_text)}",
                    f"审计结论：{paper.get('审计结论模板', '')}",
                ]
            )
            report_text = "\n".join(
                [
                    f"问题标题：{report.get('问题标题', finding.get('问题认定', ''))}",
                    f"问题描述：{report.get('问题描述模板', finding_text)}",
                    f"法规依据：{report.get('法规依据', rule.get('法规依据', ''))}",
                    f"整改建议：{report.get('整改建议模板', '')}",
                ]
            )
            conn.execute(
                """
                INSERT INTO risk_results(project_id, rulebook_id, rule_code, risk_level, score, matched_terms,
                    finding, legal_basis, working_paper_draft, report_paragraph, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    rb["id"],
                    code,
                    risk_level,
                    score,
                    "、".join(terms),
                    finding_text,
                    finding.get("法规依据") or rule.get("法规依据", ""),
                    paper_text,
                    report_text,
                    now(),
                ),
            )
            created += 1
    return created


@app.context_processor
def inject_globals():
    return {"active_rulebook": active_rulebook(), "core_sheets": CORE_SHEETS}


@app.route("/")
def index():
    with db() as conn:
        stats = {
            "rulebooks": conn.execute("SELECT COUNT(*) FROM rulebooks").fetchone()[0],
            "projects": conn.execute("SELECT COUNT(*) FROM projects").fetchone()[0],
            "files": conn.execute("SELECT COUNT(*) FROM uploaded_files").fetchone()[0],
            "risks": conn.execute("SELECT COUNT(*) FROM risk_results").fetchone()[0],
        }
    return render_template("index.html", stats=stats, default_rulebook=DEFAULT_RULEBOOK)


@app.route("/import", methods=["GET", "POST"])
def import_rules():
    if request.method == "POST":
        action = request.form.get("action")
        if action == "default":
            if not DEFAULT_RULEBOOK.exists():
                flash("未找到默认 V10 Excel 文件，请上传规则库。", "error")
            else:
                import_rulebook(DEFAULT_RULEBOOK, "工程项目招标投标审计规则引擎基础库_V10")
                flash("默认 V10 规则库已导入并设为当前规则库。", "success")
            return redirect(url_for("import_rules"))

        file = request.files.get("rulebook")
        if not file or file.filename == "":
            flash("请选择规则库 Excel 文件。", "error")
            return redirect(url_for("import_rules"))
        if Path(file.filename).suffix.lower() not in ALLOWED_RULE_EXT:
            flash("规则库仅支持 .xlsx 文件。", "error")
            return redirect(url_for("import_rules"))
        stored = DATA_DIR / f"rulebook_{uuid.uuid4().hex}.xlsx"
        file.save(stored)
        import_rulebook(stored, request.form.get("name") or file.filename)
        flash("规则库已导入并设为当前规则库。", "success")
        return redirect(url_for("import_rules"))

    with db() as conn:
        rulebooks = conn.execute("SELECT * FROM rulebooks ORDER BY id DESC").fetchall()
    return render_template("import_rules.html", rulebooks=rulebooks, default_rulebook=DEFAULT_RULEBOOK)


@app.route("/projects", methods=["GET", "POST"])
def projects():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("项目名称不能为空。", "error")
        else:
            with db() as conn:
                conn.execute(
                    "INSERT INTO projects(name, description, created_at) VALUES (?, ?, ?)",
                    (name, request.form.get("description", ""), now()),
                )
            flash("审计项目已创建。", "success")
        return redirect(url_for("projects"))
    with db() as conn:
        items = conn.execute("SELECT * FROM projects ORDER BY id DESC").fetchall()
    return render_template("projects.html", projects=items)


@app.route("/projects/<int:project_id>", methods=["GET", "POST"])
def project_detail(project_id):
    with db() as conn:
        project = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
    if not project:
        flash("项目不存在。", "error")
        return redirect(url_for("projects"))

    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("请选择上传文件。", "error")
            return redirect(url_for("project_detail", project_id=project_id))
        ext = Path(file.filename).suffix.lower()
        if ext not in ALLOWED_DOC_EXT:
            flash("资料文件支持 txt、pdf、doc、docx。", "error")
            return redirect(url_for("project_detail", project_id=project_id))
        original_name = file.filename
        stored_stem = secure_filename(Path(original_name).stem) or "upload"
        stored_name = f"{project_id}_{uuid.uuid4().hex}_{stored_stem}{ext}"
        stored_path = UPLOAD_DIR / stored_name
        file.save(stored_path)
        text = extract_text(stored_path)
        with db() as conn:
            conn.execute(
                "INSERT INTO uploaded_files(project_id, original_name, stored_name, stored_path, text_content, uploaded_at) VALUES (?, ?, ?, ?, ?, ?)",
                (project_id, original_name, stored_name, str(stored_path), text, now()),
            )
        flash("资料已上传并提取文本。", "success")
        return redirect(url_for("project_detail", project_id=project_id))

    with db() as conn:
        files = conn.execute("SELECT * FROM uploaded_files WHERE project_id=? ORDER BY id DESC", (project_id,)).fetchall()
        risks = conn.execute("SELECT * FROM risk_results WHERE project_id=? ORDER BY score DESC, id DESC LIMIT 10", (project_id,)).fetchall()
    return render_template("project_detail.html", project=project, files=files, risks=risks)


@app.route("/projects/<int:project_id>/generate", methods=["POST"])
def generate_project_risks(project_id):
    with db() as conn:
        project = conn.execute("SELECT id FROM projects WHERE id=?", (project_id,)).fetchone()
    if not project:
        flash("项目不存在。", "error")
        return redirect(url_for("projects"))

    try:
        count = generate_risks(project_id)
        flash(f"风险清单已生成，共 {count} 条匹配结果。", "success")
    except Exception as exc:
        flash(str(exc), "error")
    return redirect(url_for("risk_list", project_id=project_id))


@app.route("/projects/<int:project_id>/risks")
def risk_list(project_id):
    with db() as conn:
        project = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
        risks = conn.execute("SELECT * FROM risk_results WHERE project_id=? ORDER BY score DESC, id DESC", (project_id,)).fetchall()
    if not project:
        flash("项目不存在。", "error")
        return redirect(url_for("projects"))
    return render_template("risk_list.html", project=project, risks=risks)


@app.route("/projects/<int:project_id>/working-papers")
def working_papers(project_id):
    with db() as conn:
        project = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
        risks = conn.execute("SELECT * FROM risk_results WHERE project_id=? ORDER BY score DESC, id DESC", (project_id,)).fetchall()
    if not project:
        flash("项目不存在。", "error")
        return redirect(url_for("projects"))
    return render_template("working_papers.html", project=project, risks=risks)


@app.route("/projects/<int:project_id>/reports")
def reports(project_id):
    with db() as conn:
        project = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
        risks = conn.execute("SELECT * FROM risk_results WHERE project_id=? ORDER BY score DESC, id DESC", (project_id,)).fetchall()
    if not project:
        flash("项目不存在。", "error")
        return redirect(url_for("projects"))
    return render_template("reports.html", project=project, risks=risks)


@app.route("/rules")
def rules_index():
    rb = active_rulebook()
    sheets = []
    if rb:
        with db() as conn:
            sheets = conn.execute(
                "SELECT sheet_name, COUNT(*) AS total FROM rule_rows WHERE rulebook_id=? GROUP BY sheet_name ORDER BY sheet_name",
                (rb["id"],),
            ).fetchall()
    return render_template("rules_index.html", sheets=sheets)


@app.route("/rules/<sheet_name>")
def view_sheet(sheet_name):
    rb = active_rulebook()
    if not rb:
        flash("请先导入规则库。", "error")
        return redirect(url_for("import_rules"))
    page = max(int(request.args.get("page", 1)), 1)
    page_size = 50
    offset = (page - 1) * page_size
    with db() as conn:
        total = conn.execute(
            "SELECT COUNT(*) FROM rule_rows WHERE rulebook_id=? AND sheet_name=?", (rb["id"], sheet_name)
        ).fetchone()[0]
        rows = conn.execute(
            "SELECT row_json FROM rule_rows WHERE rulebook_id=? AND sheet_name=? ORDER BY row_index LIMIT ? OFFSET ?",
            (rb["id"], sheet_name, page_size, offset),
        ).fetchall()
    data = [json.loads(r["row_json"]) for r in rows]
    headers = list(data[0].keys()) if data else []
    return render_template("sheet.html", sheet_name=sheet_name, headers=headers, rows=data, page=page, total=total, page_size=page_size)


@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)


ensure_default_rulebook()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5002"))
    app.run(host="0.0.0.0", port=port, debug=os.environ.get("FLASK_DEBUG") == "1")
